import json
import openai
from openpyxl import Workbook , load_workbook
from tqdm import tqdm
import sys
import re
def generatePrompt(objectName):
    prompt="\
    You are an AI assistant. Your task is to annotate a given object type using your common-sense knowledge with the following values (output valid JSON):\n\
        {\"Object\":\"the given object name\",\
        \"inKitchens\":\"How likely is it for this object to appear in the kitchen or not?, return a weight, rw ∈ {0, 1, 2}, corresponding to how likely it is to appear in kitchen, the 0 indicates the object should never appear in kitchen,so ;the 1 indicates the object may appear, but is unlikely; the 2 indicates that the object appears quite often.For example, objects such as apple, bowl, bread, butterknife, chair, coffeemachine, countertop, cup, dishsponge are usually placed in the kitchen, so this field should be labeled as 2; while objects like book, bottle, cellphone, creditcard, diningtable, garbagebag may be placed in the kitchen but not often, so this field should be labeled as 1; and objects like bed, dresser, pillow are unlikely to be placed in the kitchen, so they should be labeled as 0. \",\
        \"inLivingRooms\":\"How likely is it for this object to appear in the livingroom or not?, return a weight, rw ∈ {0, 1, 2}, corresponding to how likely it is to appear in livingroom, the 0 indicates the object should never appear in living room; the 1 indicates the object may appear, but is unlikely; the 2 indicates that the object appears in livingroom quite often. For example, objects such as chair, garbagecan, watch, armchair,floorlamp, remotecontrol,sofa are usually placed in the livingroom, so this field should be labeled as 2; while objects like bowl, plate, pen, coffeetable, diningtable may be placed in the livingroom but not often, so this field should be labeled as 1; and objects like bed, winebottle, basketball, alarmclock are unlikely to be placed in the livingroom, so they should be labeled as 0. \",\
        \"inBedrooms\":\"How likely is it for this object to appear in the bedroom or not?, return a weight, rw ∈ {0, 1, 2}, corresponding to how likely it is to appear in bedroom, the 0 indicates the object should never appear in bedroom; the 1 indicates the object may appear, but is unlikely; the 2 indicates that the object appears in bedroom quite often.For example, objects such as chair, garbagecan, lightswitch, creditcard, sidetable, desklamp, labtop, pillow are usually placed in the bedroom, so this field should be labeled as 2; while objects like remotecontrol, sofa, coffeetable, basketball, teddybear may be placed in the bedroom but not often, so this field should be labeled as 1; and objects like apple, bread, knife, potato are unlikely to be placed in the bedroom, so they should be labeled as 0. \",\
        \"inBathrooms\":\"How likely is it for this object to appear in the bathroom or not?, return a weight, rw ∈ {0, 1, 2}, corresponding to how likely it is to appear in bathroom, the 0 indicates the object should never appear in bathroom; the 1 indicates the object may appear, but is unlikely; the 2 indicates that the object appears in bathroom quite often. For example, objects such as lightswitch, faucet, soapbottle, spraybottle, handtower, showerhead, soapbar, towerholder are usually placed in the bathroom, so this field should be labeled as 2; while objects like tissuebox, footstool, laundryhamper, washingmachine, clothesdryer may be placed in the bathroom but not often, so this field should be labeled as 1; and objects like bed, cd, sofa, watch, safe, desklamp are unlikely to be placed in the bathroom, so they should be labeled as 0.\",\
        \"inCorner\":\"whether this object is typically placed in the corner or not, return true or false with no explanations.This would be TRUE for a armchair, a garbagecan,a sidetable,dogbed,fridge, but FALSE for an alarmclock,a laptop, a cellphone, an apple, a pan, a cup and newspaper.\",\
        \"inMiddle\":\"whether this object is typically placed in the middle of a room or not, return true or false with no explanations.This would be FALSE for an apple, a bed, a chair, a sidetable but TRUE for toy\",\
        \"onEdge\":\"whether this object is typically placed on the edge of a room or not, return true or false with no explanations.This would be TRUE for a basketball, a bed, a coffeetable, a dogbed but FALSE for tower, a laptop, a pillow.\",\
        \"onFloor\": \"whether this object is typically placed on the floor or not, return true or false with no explanations. This would be TRUE for a piano, a bed, a sofa, a coffetable, a fridge but FALSE for a curtain, a lightswitch, a candle, a laptop, an apple and a desklamp.\",\
        \"onWall\":\"whether this object is typically placed on the wall of a room or not, return true or false with no explanations. This would be TRUE for lightswitch, handertowerholder, binds and windows,but FALSE for a bed, desklamp, tower and laptop.\",\
        \"isPickupable\":\"whether the object is pickupable or not, return true or false with no explanations.This would be TRUE for an apple, a book, cd, watch but FALSE for a bed and sofa.\",\
        \"isKinematic\":\"whether this object is kinematic or not,return true or false with no explanations. This would be TRUE for blinds,sofa but FALSE for an apple or dogbed.\",\
        \"multiplePerRoom\":\"whether there can be multiple instances of the object type in a single room, return true or false with no explanations.This would be TRUE for armchair, pen, book but FALSE for bed, microwave, soapbottle.\"\
        \"NearWindow\":\"Whether the object is typically placed near a window, return true or false with no explanations.\",\
        \"OnTable\":\"Whether the object is typically placed on a table or countertop,return true or false with no explanations.\",\
        \"InShelf\":\"Whether the object is typically placed on a shelf,return true or false with no explanations.\",\
        \"NearDoor\": \"Whether the object is typically placed near a door,return true or false with no explanations.\",\
        \"NearOutlet\": \"Whether the object is typically placed near an outlet,return true or false with no explanations.\",\
        \"InDrawer\": \"Whether the object is typically placed in a drawer,return true or false with no explanations.\",\
        \"NearWall\": \"Whether the object is typically placed near a wall,return true or false with no explanations.\",\
        \"Suspensible\": \"Whether the object can be suspended,return true or false with no explanations.\"}\
    Note 1: If this object is usually placed on the floor (i.e. \"onFloor\" is \"TRUE\"), at least one of \"inCorner\", \"inMiddle\", \"onEdge\" must be \"TRUE\" to indicate where the object is typically placed on the room floor;\
    Note 2: It is required to specify that at least one of the fields \"inKitchens\", \"inLivingRooms\", \"inBedrooms\", \"inBathrooms\" must be \"TRUE\" for the object.\
    Note 3: Ensure that the labeling results of objects conform to daily life and human common sense.\
    The user will input only the name of one object type (e.g., alarmclock). You are required to annotate this object type based on the above requirements and return a JSON-formatted string, for example:{\"Object\":\"alarmclock\",\"inKitchens\":0,\"inLivingRooms\":0,\"inBedrooms\":2,\"inBathrooms\":0,\"inCorner\":\"FALSE\",\"inMiddle\":\"FALSE\",\"onEdge\":\"FALSE\",\"onFloor\":\"FALSE\",\"onWall\":\"FALSE\",\"isPickupable\":\"TRUE\",\"isKinematic\":\"FLASE\",\"multiplePerRoom\":\"FALSE\",\"NearWindow\":\"FALSE\",\"OnTable\":\"TRUE\", \"InShelf\": \"FALSE\",\"NearDoor\":\"FALSE\",\"NearOutlet\":\"TRUE\", \"InDrawer\": \"TRUE\",\"NearWall\":\"FALSE\",\"Suspensible\": \"FALSE\"}\n\
    user:alarmclock\n\
    agent:{\"Object\":\"alarmclock\",\"inKitchens\":0,\"inLivingRooms\":0,\"inBedrooms\":2,\"inBathrooms\":0,\"inCorner\":\"FALSE\",\"inMiddle\":\"FALSE\",\"onEdge\":\"FALSE\",\"onFloor\":\"FALSE\",\"onWall\":\"FALSE\",\"isPickupable\":\"TRUE\",\"isKinematic\":\"FLASE\",\"multiplePerRoom\":\"FALSE\",\"NearWindow\":\"FALSE\",\"OnTable\":\"TRUE\", \"InShelf\": \"FALSE\",\"NearDoor\":\"FALSE\",\"NearOutlet\":\"TRUE\", \"InDrawer\": \"TRUE\",\"NearWall\":\"FALSE\",\"Suspensible\": \"FALSE\"}\n\
    user:apple\n\
    agent:{\"Object\":\"apple\",\"inKitchens\":1,\"inLivingRooms\":0,\"inBedrooms\":0,\"inBathrooms\":0,\"inCorner\":\"FALSE\",\"inMiddle\":\"FALSE\",\"onEdge\":\"FALSE\",\"onFloor\":\"FALSE\",\"onWall\":\"FALSE\",\"isPickupable\":\"TRUE\",\"isKinematic\":\"FLASE\",\"multiplePerRoom\":\"TRUE\",\"NearWindow\":\"FALSE\",\"OnTable\":\"TRUE\", \"InShelf\": \"FALSE\",\"NearDoor\":\"FALSE\",\"NearOutlet\":\"FALSE\", \"InDrawer\": \"FALSE\",\"NearWall\":\"FALSE\",\"Suspensible\": \"FALSE\"}\n\
    user:armchair\n\
    agent:{\"Object\":\"armchair\",\"inKitchens\":0,\"inLivingRooms\":2,\"inBedrooms\":1,\"inBathrooms\":0,\"inCorner\":\"TRUE\",\"inMiddle\":\"FALSE\",\"onEdge\":\"TRUE\",\"onFloor\":\"TRUE\",\"onWall\":\"FALSE\",\"isPickupable\":\"FALSE\",\"isKinematic\":\"TRUE\",\"multiplePerRoom\":\"TRUE\",\"NearWindow\":\"FALSE\",\"OnTable\":\"FALSE\", \"InShelf\": \"FALSE\",\"NearDoor\":\"FALSE\",\"NearOutlet\":\"FALSE\", \"InDrawer\": \"FALSE\",\"NearWall\":\"FALSE\",\"Suspensible\": \"FALSE\"}\n\
    user:bed\n\
    agent:{\"Object\":\"bed\",\"inKitchens\":0,\"inLivingRooms\":0,\"inBedrooms\":2,\"inBathrooms\":0,\"inCorner\":\"TRUE\",\"inMiddle\":\"FALSE\",\"onEdge\":\"FALSE\",\"onFloor\":\"TRUE\",\"onWall\":\"FALSE\",\"isPickupable\":\"FALSE\",\"isKinematic\":\"FALSE\",\"multiplePerRoom\":\"FALSE\",\"NearWindow\":\"TRUE\",\"OnTable\":\"FALSE\", \"InShelf\": \"FALSE\",\"NearDoor\":\"FALSE\",\"NearOutlet\":\"TRUE\", \"InDrawer\": \"FALSE\",\"NearWall\":\"TRUE\",\"Suspensible\": \"FALSE\"}\n\
    user:"
    return f"{prompt}{objectName}\nagent:"


if len(sys.argv) > 2:
    arg1 = sys.argv[1]
    arg2 = sys.argv[2]

    print("unlabeled objects file:", arg1)
    error_items = []
    print("Output file of object annotation:", arg2)
    api_key = 'sk-qmu3GtIMZtNYCTMm743199219bD44791BfBcDbFd9d1b3404'
    base_url = 'https://yeysai.com/v1'
    input_file=arg1
    with open(input_file, 'r') as f:
        data_require=json.load(f)
    unlabeled_objects=data_require['unlabeled_objects']
    # unlabeled_objects=["bg","french","kitchen"]

    result={"Object":"alarmclock","inKitchens":0,"inLivingRooms":0,"inBedrooms":2,"inBathrooms":0,"inCorner":"FALSE","inMiddle":"FALSE","onEdge":"FALSE","onFloor":"FALSE","onWall":"FALSE","isPickupable":"TRUE","isKinematic":"FLASE","multiplePerRoom":"FALSE","NearWindow":"FALSE","OnTable":"FALSE", "InShelf": "FALSE","NearDoor":"FALSE","NearOutlet":"TRUE", "InDrawer": "FALSE","NearWall":"TRUE","Suspensible": "FALSE"}
    keys=result.keys()
    # print(keys)
    data_list=keys
    wb=Workbook()
    ws=wb.active
    ws.delete_rows(1, ws.max_row)
    data_list=list(data_list)
    ws.append(data_list)

    tokens=0
    for i in tqdm(range(len(unlabeled_objects))):
        objectType=unlabeled_objects[i]
        # print(generatePrompt(objectName=objectType))
        client = openai.OpenAI(api_key=api_key, base_url=base_url)
        messages = [
            {"role": "user", "content": generatePrompt(objectName=objectType)}
        ]
        response = client.chat.completions.create(
            model='gpt-4', # 'gpt-3.5-turbo', 'gpt-4', 'gpt-3.5-turbo-16k', 'gpt-4-32k'
            messages=messages
        )
        tokens+=response.usage.completion_tokens
        result=response.choices[0].message.content
        try:
            result = re.search(r'\{.*\}', result, re.DOTALL).group(0)
            result = json.loads(result)
            # if(keys==result.keys()):
                # if(result["onEdge"]=="TRUE" or result["NearWall"]=="TRUE" or result["NearDoor"]=="TRUE" or result["NearWindow"]=="TRUE" or result["NearOutlet"]=="TRUE"):
                #     result["onEdge"]="TRUE"
                # if(result["onFloor"]=="TRUE" or result["Suspensible"]=="FALSE"):
                #     result["onFloor"]="TRUE"
                # if((result[""]=="FALSE" and result["InDrawer"]=="TRUE")):
                #     result["onFloor"]="TRUE"
                # if(result["onEdge"]=="TRUE" or result["NearWall"]=="TRUE" or result["NearDoor"]=="TRUE" or result["NearWindow"]=="TRUE" or result["NearOutlet"]=="TRUE"):
                    # result["onEdge"]="TRUE"
                # print(result)
        except json.JSONDecodeError as e:
            print(f"JSON 解码异常: {e}")
            error_items.append(objectType)
            print(result)
            continue
        data_list=[]
        for cell in ws[1]:
            type_name=cell.value
            # print(type_name)
            if type_name in result:
                data_list.append(result[type_name])
            else:
                data_list.append(None)
        ws.append(data_list)
    print("tokens:",tokens)
    wb.save(arg2)
    if len(error_items)>0 and error_items is not None:
        with open('error_items_object.txt', 'w') as f:
            for item in error_items:
                # write each item on a new line
                f.write("%s\n" % item)
else:
    print("Not enough arguments provided.")
