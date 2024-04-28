import pandas as pd
import json
import openai
from openpyxl import Workbook , load_workbook
from tqdm import tqdm
import sys
import re
def Prompt_recp(objectName):
    prompt="\
    You are an AI assistant.The user will provide the name of an object.Your task is to annotate whether the object is a receptacle or not. In other words, can things be placed on this object? For example, a bowl is a receptacle, but a pencil is not.A desk is a receptacle,but water is not.\n\
    The user will input only the name of one object type (e.g., alarmclock). You are required to annotate this object type based on the above requirements and return a JSON-formatted string, for example:{\"Object\":\"alarmclock\",\"isReceptacle\":\"FALSE\"}\n\
    Please provide the correct answer based on common sense. \
    usr:alarmclock\n\
    agent:{\"Object\":\"alarmclock\",\"isReceptacle\":\"FALSE\"}\n\
    usr:apple\n\
    agent:{\"Object\":\"apple\",\"isReceptacle\":\"FALSE\"}\n\
    usr:armchair\n\
    agent:{\"Object\":\"armchair\",\"isReceptacle\":\"TRUE\"}\n\
    usr:bed\n\
    agent:{\"Object\":\"bed\",\"isReceptacle\":\"TRUE\"}\n\
    usr:"
    return f"{prompt}{objectName}\nagent:"
if len(sys.argv) > 2:
    arg1 = sys.argv[1]
    arg2 = sys.argv[2]
    print("unlabeled objects file:", arg1)
    print("Output file about object is a receptacle or not:", arg2)
    api_key = 'sk-qmu3GtIMZtNYCTMm743199219bD44791BfBcDbFd9d1b3404'
    base_url = 'https://yeysai.com/v1'
    input_file=arg1
    error_items=[]
    with open(input_file, 'r') as f:
        data_require=json.load(f)
    unlabeled_objects=data_require['unlabeled_objects']
    objects_name=unlabeled_objects

    result={"Object":"alarmclock","isReceptacle":"FALSE"}
    keys=result.keys()
    # print(keys)
    data_list=keys
    wb=Workbook()
    ws=wb.active
    ws.delete_rows(1, ws.max_row)
    data_list=list(data_list)
    ws.append(data_list)


    for i in tqdm(range(len(objects_name))):
        objectType=objects_name[i]
        client = openai.OpenAI(api_key=api_key, base_url=base_url)
        messages = [
            {"role": "user", "content": Prompt_recp(objectName=objectType)}
        ]
        response = client.chat.completions.create(
            model='gpt-4', # 'gpt-3.5-turbo', 'gpt-4', 'gpt-3.5-turbo-16k', 'gpt-4-32k'
            messages=messages
        )
        # print(response.choices[0].message.content)
        result=response.choices[0].message.content
        try:
            result = re.search(r'\{.*\}', result, re.DOTALL).group(0)
            result = json.loads(result)
        except json.JSONDecodeError as e:
            error_items.append(objectType)
            print(f"JSON 解码异常: {e}")

            continue
        data_list=[]
        # ws.append(["Object","Kitchen","Living Room","Bedroom","Bathroom","Corner","Middle","On Edge","On Floor","On Wall","Pickupable","Kinematic","Multiple Per Room"])
        for cell in ws[1]:
            type_name=cell.value
            # print(type_name)
            if type_name in result:
                data_list.append(result[type_name])
            else:
                data_list.append(None)
        ws.append(data_list)
    wb.save(arg2)
    if len(error_items)>0 and error_items is not None:
        with open('error_items_receptacle.txt', 'w') as f:
            for item in error_items:
                # write each item on a new line
                f.write("%s\n" % item)

        